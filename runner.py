#!/usr/bin/python
"""runner.py: Includes the main class for controlling the flow of the game. Input move updates and the bot displays
information about its and others' turns."""
__author__ = "Christopher Lehman"
__email__ = "lehman40@purdue.edu"

from cluechecker import check_space_with_clue, get_clue_dict, check_all_spaces_with_clue, check_all_clues_with_space, get_start_time, get_players, get_clues, get_board_config
import itertools
import random
import datetime
from openpyxl import load_workbook
import time
import traceback

def play():
    players = get_players()

    names = []
    for i in range(2, players + 1):
        names.append('Player ' + str(i))

    # mode = int(input('Which mode? \n1 - Normal \n2 - Advanced\n'))
    mode = 2
    # ideal_prop represents the proportion of clues the bot would learn about a player's clue when questioning. We want
    # this around 30% (varies with player number) so that the player is likely to place a disc. Or if they place a cube,
    # we learn a lot about their clue (but have to place one of our own)
    if players == 5:
        ideal_prop = .25
    elif players == 4:
        ideal_prop = .3
    else:
        ideal_prop = .35

    try:
        my_clue_text = get_clues()[0]
        print(my_clue_text)
        my_clue = get_clue_dict()[my_clue_text]
        spots_to_avoid = []

        orig_clue_dict = dict((v, k) for k, v in get_clue_dict().items())

        # which clues other players think it could have
        my_possible = list(range(1, 49) if mode == 2 else list(range(1, 24)))
        # which spaces have pieces on them, should avoid for asking
        non_satisfied_spaces = list(set(range(1, 109)) - set(check_all_spaces_with_clue(my_clue)))
        inital_spaces_out = len(non_satisfied_spaces)

        # which clues other players could have
        others_remaining = [list(range(1, 49) if mode == 2 else list(range(1, 24))) for _ in range(players - 1)]

        # If its clue is a two terrain one, can eliminate some clues from others's possibilities
        if my_clue < 11:
            terrains = (my_clue_text.split()[4], my_clue_text.split()[6])
            for i in range(1, 11):
                if terrains[0] not in orig_clue_dict[i] and terrains[1] not in orig_clue_dict[i]:
                    for o in others_remaining:
                        o.remove(i)
        # If its clue is a 'not', the inverse is not possible for the others
        if my_clue > 24:
            for o in others_remaining:
                o.remove(my_clue - 24)
        # The others cannot have its clue
        for o in others_remaining:
            o.remove(my_clue)

        turn = 1
    except:
        traceback.print_exc()
        workbook = load_workbook(filename="Cryptid_Logs.xlsx")
        sheet = workbook.active
        for cell in sheet["A"]:
            if cell.value is None:
                next_row = cell.row
                break
        else:
            next_row = cell.row + 1
        sheet['A' + str(next_row)] = 'false'
        sheet['B' + str(next_row)] = players
    while True:
        if turn > players:
            turn = 1
        if turn == 1:  # Bot is player 1
            print("It's my turn!")
            all_possible_spaces = {}

            # If the total combination of others possible clues is reasonably small. Compute all the combinations that
            # only lead to one space, otherwise, that combination is invalid
            print(total_comb(others_remaining))
            if total_comb(others_remaining) < 130000:
                others_possibilities = [set() for _ in range(players - 1)]
                possibles_cumul = []
                for possible in itertools.product(*others_remaining):
                    working_spaces = set.intersection(set(check_all_spaces_with_clue(my_clue)), *[set(check_all_spaces_with_clue(p)) for p in possible])
                    # Check if just one space works and this possible set of clues are all distinct
                    if len(working_spaces) == 1 and len(set(possible)) == len(possible):
                        space = working_spaces.pop()
                        if set(possible) not in possibles_cumul:
                            possibles_cumul.append(set(possible))
                            if space in all_possible_spaces:
                                all_possible_spaces[space] += 1
                            else:
                                all_possible_spaces[space] = 1
                        for i, p in enumerate(possible):
                            others_possibilities[i].add(p)
                # Update the possible remaining with only the clues that led to once space
                others_remaining = [list(o) for o in others_possibilities]
                print(others_remaining)
                print(all_possible_spaces)

            all_possible_spaces_ordered = sorted(all_possible_spaces.items(), key=lambda kv: kv[1], reverse=True)
            workbook = load_workbook(filename="Cryptid_Logs.xlsx")
            sheet = workbook.active
            for cell in sheet["A"]:
                if cell.value is None:
                    next_row = cell.row
                    break
            else:
                next_row = cell.row + 1

            # Check if there was only one possible space or a search is reasonable (see method check_to_search below)
            if len(all_possible_spaces) == 1 or check_to_search(all_possible_spaces, others_remaining):
                if len(all_possible_spaces) == 1:
                    space = list(all_possible_spaces.keys())[0]
                else:
                    # determine the best space to search of the possible ones based on how much it reveals from its clue
                    # by placing a piece and the likelihood of the player to the left placing a disc (means you get
                    # extra information)
                    candidates = []
                    for space in all_possible_spaces:
                        if len(check_all_clues_with_space(space, others_remaining[0])) / len(others_remaining[0]) > .49:
                            candidates.append(space)
                    m = list(all_possible_spaces.values())[0]
                    mk = list(all_possible_spaces.keys())[0]
                    for c in candidates:
                        if all_possible_spaces[c] > m:
                            m = all_possible_spaces[c]
                            mk = c
                    space = mk
                row = (space - 1) // 12
                col = (space - 1) % 12
                # Make the search
                print("I think it's space " + str(space) + " at row " + str(row) + " and column " + str(col))
                # Loop through only players or until a cube is placed
                result_space = max(set.intersection(*[set(check_all_spaces_with_clue(c)) for c in [get_clue_dict()[clue] for clue in get_clues()]]))

                sheet['A' + str(next_row)] = 'true'
                sheet['B' + str(next_row)] = get_board_config()
                sheet['C' + str(next_row)] = players
                sheet['D' + str(next_row)] = str(round(time.time() - get_start_time(), 2))
                sheet['E' + str(next_row)] = my_clue_text
                sheet['F' + str(next_row)] = inital_spaces_out
                sheet['G' + str(next_row)] = len(all_possible_spaces)
                sheet['H' + str(next_row)] = 'true'
                sheet['I' + str(next_row)] = 'true' if result_space == space else 'false'
                sheet['J' + str(next_row)] = result_space
                col_ind = 11
                print(all_possible_spaces_ordered)
                for space_num, freq in all_possible_spaces_ordered:
                    print(space_num, freq)
                    sheet[colnum_string(col_ind) + str(next_row)] = space_num
                    col_ind += 1
                    sheet[colnum_string(col_ind) + str(next_row)] = freq
                    col_ind += 1
                    if col_ind > 40:
                        break

                workbook.save(filename="Cryptid_Logs.xlsx")
                return
            else:
                result_space = max(set.intersection(
                    *[set(check_all_spaces_with_clue(c)) for c in [get_clue_dict()[clue] for clue in get_clues()]]))

                sheet['A' + str(next_row)] = 'true'
                sheet['B' + str(next_row)] = get_board_config()
                sheet['C' + str(next_row)] = players
                sheet['D' + str(next_row)] = str(round(time.time() - get_start_time(), 2))
                sheet['E' + str(next_row)] = my_clue_text
                sheet['F' + str(next_row)] = inital_spaces_out
                sheet['G' + str(next_row)] = len(all_possible_spaces)
                sheet['H' + str(next_row)] = 'false'
                sheet['J' + str(next_row)] = result_space
                col_ind = 11
                for space_num, freq in all_possible_spaces_ordered:
                    sheet[colnum_string(col_ind) + str(next_row)] = space_num
                    col_ind += 1
                    sheet[colnum_string(col_ind) + str(next_row)] = freq
                    col_ind += 1
                    if col_ind > 40:
                        break

                workbook.save(filename="Cryptid_Logs.xlsx")
                return
                # Not enough info to search, so will question a player
                lengths = [len(o) for o in others_remaining]
                # Asks player with most possible remaining clues (decides randomly in a tie)
                players_to_ask = []
                for i, length in enumerate(lengths):
                    if length == max(lengths):
                        players_to_ask.append(i)
                player_to_ask = random.choice(players_to_ask)
                proportions = []
                # For every space, track the proportion of remaining clues the targeted player would reveal if a disc
                # was placed
                spaces = []
                for space in range(1, 109):
                    row = (space - 1) // 12
                    col = (space - 1) % 12
                    if (row, col) not in spots_to_avoid:
                        proportions.append(1 - (len(check_all_clues_with_space(space, others_remaining[player_to_ask]))
                                                / len(others_remaining[player_to_ask])))
                        spaces.append(space)
                # Reset to how close proportions are to the ideal that was set
                proportions = [abs(p - ideal_prop) for p in proportions]
                candidates = []
                for i, p in enumerate(proportions):
                    if p == min(proportions):
                        candidates.append(spaces[i])
                space = random.choice(candidates)
                # Chose a random space from all the closest ones (more than one if a tie)
                row = (space - 1) // 12
                col = (space - 1) % 12

                # Prepare the question
                found = int(input(names[player_to_ask] + ", could it be at row " + str(row) + " and column " + str(col) + "?\n1 - Yes \n2 - No\n"))

                spots_to_avoid.append((row, col))
                if found == 1:
                    # It could be there, update info
                    others_remaining[player_to_ask] = update_disc(others_remaining[player_to_ask], row, col)
                else:
                    # It can't be there, update info and prepare to place a cube
                    others_remaining[player_to_ask] = update_cube(others_remaining[player_to_ask], row, col)
                    show_cube = []
                    spaces = []
                    for space in non_satisfied_spaces:
                        row = (space - 1) // 12
                        col = (space - 1) % 12
                        if (row, col) not in spots_to_avoid:
                            # track how many clues placing a cube on each space would reveal
                            show_cube.append(len(check_all_clues_with_space(space, my_possible)))
                            spaces.append(space)
                    candidates = []
                    for i, spot in enumerate(show_cube):
                        if spot == min(show_cube):
                            candidates.append(i)
                    space_to_place = spaces[random.choice(candidates)]

                    row = (space_to_place - 1) // 12
                    col = (space_to_place - 1) % 12
                    spots_to_avoid.append((row, col))
                    print("Ok. For me it can't be at row " + str(row) + " and column " + str(col) + ". Could you place a cube for me as I have no arms?")
                    my_possible = update_cube(my_possible, row, col)

        else:
            print("It's " + names[turn - 2] + "'s turn")
            move = int(input('Which move type did they take? \n1 - Question \n2 - Search\n'))
            if move == 1:
                target = input("Who was asked? (Type the name of the player. If it's me, the bot, type anything else) ").title()
                asked = names.index(target) + 2 if target in names else 1
                if asked == 1:
                    loc = input("Where was I asked? ")
                    row = int(loc.split()[0])
                    col = int(loc.split()[1])
                    spots_to_avoid.append((row, col))
                    print("Let me check my clue...")
                    # Checks its clue for that space
                    print("It could be there!" if check_space_with_clue(row * 12 + col + 1, my_clue) else "No such luck!")
                    if check_space_with_clue(row * 12 + col + 1, my_clue):
                        my_possible = update_disc(my_possible, row, col)
                    else:
                        my_possible = update_cube(my_possible, row, col)
                        # Can't be there so they other player must place a cube
                        loc = input("Where did " + names[turn - 2] + " place the cube? ")
                        row = int(loc.split()[0])
                        col = int(loc.split()[1])
                        others_remaining[turn - 2] = update_cube(others_remaining[turn - 2], row, col)
                else:
                    loc = input("Where were they asked? ")
                    row = int(loc.split()[0])
                    col = int(loc.split()[1])
                    spots_to_avoid.append((row, col))
                    found = int(input('Could it be there?\n1 - Yes \n2 - No\n'))
                    if found == 1:
                        others_remaining[asked - 2] = update_disc(others_remaining[asked - 2], row, col)
                    else:
                        others_remaining[asked - 2] = update_cube(others_remaining[asked - 2], row, col)
                        loc = input("Where did " + names[turn - 2] + " place the cube? ")
                        row = int(loc.split()[0])
                        col = int(loc.split()[1])
                        spots_to_avoid.append((row, col))
                        others_remaining[turn - 2] = update_cube(others_remaining[turn - 2], row, col)
            else:
                # A player searched a specific place
                loc = input("Where was searched? ")
                row = int(loc.split()[0])
                col = int(loc.split()[1])
                spots_to_avoid.append((row, col))
                if fp < len(lines):
                    real_loc = lines[fp]
                    fp += 1
                else:
                    real_loc = input("Put the other location if the disc had to be placed elsewhere, otherwise enter: ")
                fw.write(real_loc + '\n')
                if len(real_loc) > 1:
                    rrow = int(real_loc.split()[0])
                    rcol = int(real_loc.split()[1])
                    spots_to_avoid.append((rrow, rcol))
                    others_remaining[turn - 2] = update_disc(others_remaining[turn - 2], rrow, rcol)
                else:
                    others_remaining[turn - 2] = update_disc(others_remaining[turn - 2], row, col)
                current = turn
                for i in range(1, players):
                    # Loop through all other players
                    current += 1
                    if current > players:
                        current = 1
                    if current == 1:
                        print("Let me check if it could be there...")
                        print("It could be there!" if check_space_with_clue(row * 12 + col + 1,
                                                                            my_clue) else "No such luck!")
                        if check_space_with_clue(row * 12 + col + 1, my_clue):
                            my_possible = update_disc(my_possible, row, col)
                        else:
                            # Bot placed the cube
                            my_possible = update_cube(my_possible, row, col)
                            break
                        if current == turn - 1:
                            # Everyone placed a disc
                            print("I am defeated, well played.")
                            print("My clue was:", my_clue_text)
                            new_file = open(
                                "Game_Records\\Game-" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M").replace(':', '-') + ".txt",
                                "w")
                            with open("lastGameWrite.txt", "r") as f:
                                new_file.write(f.read())
                            new_file.close()
                            print("Game saved at:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
                            return
                    else:
                        if fp < len(lines):
                            found = int(lines[fp])
                            fp += 1
                        else:
                            found = int(input('Did ' + names[current - 2] + ' place a disc?\n1 - Yes \n2 - No\n'))
                        fw.write(str(found) + '\n')
                        if found == 1:
                            others_remaining[current - 2] = update_disc(others_remaining[current - 2], row, col)
                            if current == turn - 1:
                                # Everyone placed a disc
                                print("I am defeated, well played.")
                                print("My clue was:", my_clue_text)
                                new_file = open("Game_Records\\Game-" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M").replace(':', '-') + ".txt",
                                                "w")
                                with open("lastGameWrite.txt", "r") as f:
                                    new_file.write(f.read())
                                new_file.close()
                                print("Game saved at:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
                                return
                        else:
                            # Cube was placed
                            others_remaining[current - 2] = update_cube(others_remaining[current - 2], row, col)
                            break
        turn += 1


# Convert column number to excel letter
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


# Calculate the total possible combinations of others' remaining clues
def total_comb(others_remaining):
    t = 1
    for o in others_remaining:
        t *= len(o)
    return t


# Updates a list of clues by removing clues that don't work with the space a disc was placed
def update_disc(remaining, r, c):
    return [clue for clue in remaining if check_space_with_clue(r * 12 + c + 1, clue)]


# Updates a list of clues by removing clues that do work with the space a cube was placed
def update_cube(remaining, r, c):
    return [clue for clue in remaining if not check_space_with_clue(r * 12 + c + 1, clue)]


# Check if a search is warranted
def check_to_search(all_possible, others_remaining):
    # Not worth it if there are too many spaces that could work
    if len(all_possible) > 4:
        return False
    # Check if there is at least one space of the possible that there is a good enough chance the player to the left
    # places a disc (65%)
    for space in all_possible:
        # print(len(check_all_clues_with_space(space, my_possible)) / len(my_possible))
        # print(len(check_all_clues_with_space(space, others_remaining[0])) / len(others_remaining[0]))
        if len(check_all_clues_with_space(space, others_remaining[0])) / len(others_remaining[0]) > .49:
            return True
    print('CASE FOUND')
    print(get_clues())
    print(all_possible)
    print(others_remaining)
    for space in all_possible:
        print(space)
        print(check_all_clues_with_space(space, others_remaining[0]))

    return False


play()
